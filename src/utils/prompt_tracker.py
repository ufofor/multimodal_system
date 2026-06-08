"""Tracks prompt candidates and scores across sessions to ~/Documents/prompt_archive.xlsx."""
import os
from dataclasses import dataclass
from datetime import date
from pathlib import Path

import openpyxl

ARCHIVE_PATH = Path.home() / "Documents" / "prompt_archive.xlsx"
HEADERS = [
    "Date", "Session_Type", "Task", "Prompt_ID", "Title",
    "Efficiency", "Accuracy", "Relevance", "Autonomy", "Output_Quality",
    "Total", "Winner",
]


@dataclass
class PromptCandidate:
    prompt_id: str        # e.g. "P1", "P2"
    title: str            # short label
    text: str             # full prompt text
    efficiency: int = 0
    accuracy: int = 0
    relevance: int = 0
    autonomy: int = 0
    output_quality: int = 0

    @property
    def total(self) -> int:
        return self.efficiency + self.accuracy + self.relevance + self.autonomy + self.output_quality


def save_session(
    task: str,
    candidates: list[PromptCandidate],
    winner_id: str,
    session_type: str = "rag",  # "rag" or "claude_code"
) -> None:
    """Append one session's prompt candidates to the Excel archive."""
    if ARCHIVE_PATH.exists():
        wb = openpyxl.load_workbook(ARCHIVE_PATH)
        ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(HEADERS)
        _style_header(ws)

    today = date.today().isoformat()
    for c in candidates:
        ws.append([
            today,
            session_type,
            task,
            c.prompt_id,
            c.title,
            c.efficiency,
            c.accuracy,
            c.relevance,
            c.autonomy,
            c.output_quality,
            c.total,
            "Yes" if c.prompt_id == winner_id else "No",
        ])

    wb.save(ARCHIVE_PATH)


def _style_header(ws) -> None:
    from openpyxl.styles import Font, PatternFill
    fill = PatternFill("solid", fgColor="1F3564")
    font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
